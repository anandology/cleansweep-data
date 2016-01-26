# -*- coding: utf-8 -*-
from openpyxl import load_workbook  # pip install openpyxl

'''
This was made for an excel sheet like this:
Sector Number	Assembly Number & Name	                           Zone No.	Zone Name
Gurdaspur-01	1-Sujanpur, 2-Bhoa (SC), 3-Pathankot	             1	     Gurdaspur
Gurdaspur-02	4-Gurdaspur, 5-Dina Nagar (SC), 6-Qadian           Null      Null
Gurdaspur-03	7-Batala, 9-Fatehgarh Churian, 10-Dera Baba Nanak   Null      Null
'''


wb = load_workbook(filename = 'data.xlsx')  # Excel file
sheet = wb.get_sheet_by_name(name = 'Sheet1')  # Default sheet name

# Columns
sector_clmn = 0
ac_clmn = 1
zone_number_clmn = 2
zone_name_clmn = 3

state_short_name = "PB"
zone_short_name = "Z"
sector_short_name = "S"
ac_short_name = "AC"

zone_file = open("data/PB/1-zone.txt", "w")
sector_file = open("data/PB/2-sector.txt", "w")
ac_file = open("data/PB/3-ac.txt", "w")

zone_number, zone_name, zone_key = 0, "", ""
for i, row in enumerate(sheet.iter_rows()):
    if i == 0:  # Skip heading
        continue
    zone_name_value = row[zone_name_clmn].internal_value
    if (zone_name_value):
        char = "A"
        zone_name = zone_name_value
        zone_number = int(row[zone_number_clmn].internal_value)
        key = zone_short_name + "%02d" % zone_number
        zone_key = state_short_name + "/" + key
        zone = state_short_name + "\t" + zone_key + "\t" + key + " - " + zone_name
        zone_file.write(zone + "\n")

    sector_name, sector_number = str(row[sector_clmn].internal_value).split("-")
    key = sector_short_name + "%02d" % int(sector_number)
    sector_key = zone_key + "/" + key
    sector = zone_key + "\t" + sector_key + "\t" + key + " - " + sector_name + " " + char
    sector_file.write(sector + "\n")
    char = chr(ord(char) + 1) # Don't know how else to do it

    # 3 AC per row
    acs = unicode(row[ac_clmn].internal_value).encode('ascii', 'ignore').split(",")
    for ac in acs:
        ac = ac.strip()
        ac_number, ac_name = ac.split("-")
        key = ac_short_name + "%03d" % int(ac_number)
        ac_key = sector_key + "/" + key
        ac = sector_key + "\t" + ac_key + "\t" + key + " - " + ac_name
        ac_file.write(ac + "\n")

zone_file.close()
sector_file.close()
ac_file.close()
